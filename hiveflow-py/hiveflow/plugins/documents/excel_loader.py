"""Excel document loader using openpyxl."""

from pathlib import Path

from hiveflow.plugins.documents import Document, DocumentLoaderPlugin


class ExcelLoader(DocumentLoaderPlugin):
    """Loader for Excel files using openpyxl."""

    @property
    def plugin_id(self) -> str:
        return "excel"

    @property
    def description(self) -> str:
        return "Excel file loader using openpyxl"

    @property
    def supported_extensions(self) -> list[str]:
        return [".xlsx", ".xls"]

    async def load(self, file_path: str | Path) -> Document:
        try:
            import openpyxl
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required for Excel support. Install with: pip install openpyxl"
            ) from exc

        path = Path(file_path)
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        sheets: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[str] = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(c for c in cells):
                    rows.append("\t".join(cells))
            if rows:
                sheets.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))

        wb.close()
        content = "\n\n".join(sheets)
        return Document(
            content=content,
            source=str(path),
            metadata={
                "filename": path.name,
                "size_bytes": path.stat().st_size,
                "sheet_count": len(sheets),
            },
        )
